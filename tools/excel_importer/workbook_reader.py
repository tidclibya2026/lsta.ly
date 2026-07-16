from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(f"الملف غير موجود: {path}")

    if path.suffix.lower() != ".xlsx":
        raise ValueError("الأداة تدعم ملفات XLSX فقط في هذه المرحلة.")

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    result: dict[str, list[dict[str, Any]]] = {}

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            result[worksheet.title] = []
            continue

        headers = [
            str(value).strip() if value is not None else f"column_{index + 1}"
            for index, value in enumerate(rows[0])
        ]

        sheet_records: list[dict[str, Any]] = []

        for row in rows[1:]:
            record = {
                headers[index]: value
                for index, value in enumerate(row)
                if index < len(headers)
            }

            if any(value not in (None, "") for value in record.values()):
                sheet_records.append(record)

        result[worksheet.title] = sheet_records

    workbook.close()
    return result