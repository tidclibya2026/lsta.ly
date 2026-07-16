from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tools.merge_engine.models import SourceRecord


def read_configured_excel(path: Path, config: dict[str, Any]) -> list[SourceRecord]:
    """Read an Excel layer using field names declared in its intake configuration."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[config["excel_sheet"]]
    rows = ws.iter_rows(values_only=True)
    headers: list[str] = []
    for row in rows:
        candidate = [str(value).strip() if value is not None else "" for value in row]
        if config["excel_id_field"] in candidate:
            headers = candidate
            break
    if not headers:
        wb.close()
        raise ValueError(f"Header row containing {config['excel_id_field']} was not found")
    index = {name: position for position, name in enumerate(headers)}
    def value(row: tuple[Any, ...], field: str | None) -> Any:
        position = index.get(field or "")
        return row[position] if position is not None and position < len(row) else None
    result=[]
    for row in rows:
        if not any(item not in (None, "") for item in row):
            continue
        source_id=str(value(row,config["excel_id_field"]) or "").strip()
        name_ar=str(value(row,config["name_ar_field"]) or "").strip()
        if not source_id or not name_ar:
            continue
        result.append(SourceRecord(source_type="excel",source_id=source_id,name_ar=name_ar,name_en=str(value(row,config.get("name_en_field")) or "").strip() or None,latitude=_to_float(value(row,config.get("latitude_field"))),longitude=_to_float(value(row,config.get("longitude_field"))),municipality=str(value(row,config.get("municipality_field")) or "").strip() or None,category_code=str(value(row,config.get("category_field")) or "").strip() or None,description=None,source_reference=path.name,properties={headers[i]:row[i] for i in range(min(len(headers),len(row))) if headers[i] in set(config.get("business_attribute_fields",[]))|set(config.get("required_fields",[]))}))
    wb.close()
    return result


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def read_hotels_excel(path: Path, sheet_name: str = "الفنادق") -> list[SourceRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    index = {name: idx for idx, name in enumerate(headers)}
    result: list[SourceRecord] = []
    for row in rows:
        if not any(v not in (None, "") for v in row):
            continue
        def get(name: str) -> Any:
            idx = index.get(name)
            return row[idx] if idx is not None and idx < len(row) else None
        result.append(SourceRecord(
            source_type="excel",
            source_id=str(get("Master_ID") or "").strip(),
            name_ar=str(get("اسم الفندق") or "").strip(),
            name_en=str(get("الاسم الإنجليزي") or "").strip() or None,
            latitude=_to_float(get("خط العرض")),
            longitude=_to_float(get("خط الطول")),
            municipality=str(get("البلدية") or "").strip() or None,
            category_code=str(get("كود التصنيف") or "").strip() or None,
            description=str(get("الوصف المختصر") or "").strip() or None,
            source_reference=str(get("ملفات المصدر") or "").strip() or None,
            properties={headers[i]: row[i] for i in range(min(len(headers), len(row)))},
        ))
    wb.close()
    return result



def _validate_geojson_path(path: Path) -> None:
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(
            "ملف GeoJSON غير موجود.\n"
            f"المسار المطلوب: {path}\n"
            "استخرج طبقة الفنادق من أحدث ملف KML/KMZ أولًا، "
            "ثم استخدم المسار الفعلي للملف الناتج."
        )

    if path.suffix.lower() not in {".geojson", ".json"}:
        raise ValueError(
            f"صيغة غير مدعومة لملف البيانات الجغرافية: {path.suffix}"
        )

def read_geojson(path: Path) -> list[SourceRecord]:
    _validate_geojson_path(path)
    data = json.loads(path.read_text(encoding="utf-8"))
    features = data.get("features", []) if isinstance(data, dict) else data
    result: list[SourceRecord] = []
    for idx, feature in enumerate(features, start=1):
        props = feature.get("properties") or {}
        geometry = feature.get("geometry") or {}
        coords = geometry.get("coordinates") or []
        lon = lat = None
        if geometry.get("type") == "Point" and isinstance(coords, list) and len(coords) >= 2:
            lon, lat = _to_float(coords[0]), _to_float(coords[1])
        source_id = str(props.get("feature_id") or props.get("id") or f"KML-{idx:06d}")
        name = str(props.get("name_ar") or props.get("name") or "").strip()
        result.append(SourceRecord(
            source_type="kml",
            source_id=source_id,
            name_ar=name,
            name_en=str(props.get("name_en") or "").strip() or None,
            latitude=lat,
            longitude=lon,
            municipality=str(props.get("municipality") or "").strip() or None,
            category_code=str(props.get("category_code") or props.get("category") or "").strip() or None,
            description=str(props.get("description_text") or props.get("description") or "").strip() or None,
            source_reference=str(props.get("source_file") or path.name),
            properties=props,
        ))
    return result
